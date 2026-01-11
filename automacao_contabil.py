import re
import sys
import tkinter as tk
from tkinter import filedialog
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook


def run_automation(file_path=None, action_delay=0.5):
    """Executa a automação usando o arquivo especificado.
    Se `file_path` for None, abre um diálogo para seleção (compatibilidade direta).

    Parâmetros de controle de velocidade:
    - action_delay (float): atraso em segundos aplicado explicitamente após várias ações
      através de `page.wait_for_timeout(int(action_delay * 1000))`. Defina `0` para
      desativar os waits explícitos entre ações.
    - slow_mo (int, milissegundos): atraso global do Playwright configurado em
      `browser.launch(...)` abaixo. `slow_mo` adiciona uma espera interna entre ações;
      para acelerar a execução, reduza ou remova `slow_mo`.
    """
    # seleção compatível com uso direto do script
    if not file_path:
        root = tk.Tk()
        root.withdraw()
        file_path = filedialog.askopenfilename(title="Selecione a planilha de lançamentos", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            print("Nenhum arquivo selecionado. Encerrando.")
            return

    with sync_playwright() as p:
        # slow_mo (ms) é um atraso global do Playwright aplicado a todas as ações.
        # - Atualmente definido como 2000 (2s). Para executar mais rápido, use 0.
        # - `action_delay` (acima) controla atrasos por-ação; ajuste conforme desejado.
        browser = p.chromium.launch(headless=False, )
        context = browser.new_context(
            viewport={"width": 1920, "height": 1080},
            accept_downloads=True,
            locale="pt-BR",
        )
        page = context.new_page()
        page.goto('https://simulacontabil.netlify.app/')
        page.wait_for_timeout(int(action_delay * 1000))

        # Digitando e-mail e senha (mantido como estava)
        page.get_by_test_id("email-input").fill("admin@simulacontabil.com.br")
        page.wait_for_timeout(int(action_delay * 1000))
        page.get_by_test_id("password-input").fill("admin")
        page.wait_for_timeout(int(action_delay * 1000))
        page.get_by_test_id("login-button").click()
        page.wait_for_timeout(int(action_delay * 1000))

        # - clicando na aba de lançamentos
        page.get_by_test_id("nav-lançamentos").click()
        page.wait_for_timeout(int(action_delay * 1000))

        # Lendo dados da planilha
        planilha = load_workbook(filename=file_path)
        pagina = planilha["Lançamentos"]

        for linha in pagina.iter_rows(min_row=2, values_only=True):
            descricao = linha[0]
            valor = linha[1]
            data = linha[2]
            tipo = linha[3]
            categoria = linha[4]
            status = linha[5]

            page.get_by_test_id("btn-new-transaction").click()
            page.wait_for_timeout(int(action_delay * 1000))

            page.get_by_test_id("input-description").fill((str(descricao)))
            page.wait_for_timeout(int(action_delay * 1000))
            page.get_by_test_id("input-amount").fill(str(valor))
            page.wait_for_timeout(int(action_delay * 1000))
            page.get_by_test_id("input-date").fill(str(data))
            page.wait_for_timeout(int(action_delay * 1000))

            if tipo == "Receita":
                page.get_by_test_id("select-type").select_option("RECEITA")
            else:
                page.get_by_test_id("select-type").select_option("DESPESA")
            page.wait_for_timeout(int(action_delay * 1000))

            page.get_by_test_id("select-category").select_option(categoria)
            page.wait_for_timeout(int(action_delay * 1000))

            page.locator("div").filter(has_text=re.compile(r"^StatusPendentePagoAtrasado$")).get_by_role("combobox").select_option(status.upper())
            page.wait_for_timeout(int(action_delay * 1000))

            page.get_by_test_id("btn-save").click()
            page.wait_for_timeout(int(action_delay * 1000))

        try:
            browser.close()
        except Exception:
            pass
        else:
            # confirmação para execuções em modo console
            print("Relatório incluído com sucesso.")


if __name__ == "__main__":
    # comportamento antigo ao executar diretamente: abre diálogo para selecionar arquivo
    run_automation()